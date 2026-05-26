#!/usr/bin/env python3
"""
Экспорт jobs_master.csv в xlsx и интерактивный HTML.

Стиль xlsx:
- Тонкие границы, чёрные заголовки на сером фоне
- Подсветка дубликатов (is_canonical=0 → серый фон строки)
- Цвета размера: Big=зелёный фон, Middle=жёлтый, Small=без фона
- Авто-фильтры, freeze pane

HTML:
- DataTables.js (CDN) — фильтры по сегменту, размеру, источнику
- Поиск по тексту
- Цветовая маркировка дубликатов
"""

import csv
from pathlib import Path
import sys

# Проверка openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("openpyxl не установлен. Устанавливаю…")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

OUT = Path("/Users/romansemenov/Desktop/Cloud Cod/Knowledge/jobs_master_table")
CSV_RAW = OUT / "jobs_master.csv"
XLSX = OUT / "jobs_master.xlsx"
HTML = OUT / "jobs_master.html"

# Читаем CSV
with open(CSV_RAW, encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

cols = list(rows[0].keys()) if rows else []

# Удобные заголовки для UI
HEADERS_RU = {
    "job_size": "Размер",
    "business_segment": "Бизнес-сегмент",
    "job_segment": "Джоб-сегмент",
    "want_short": "Название джоба / Хочу",
    "context_when": "Когда (контекст)",
    "want_result": "Хочу получить (результат)",
    "so_that": "Чтобы (цель/эффект)",
    "importance": "Важность",
    "satisfaction": "Удовлетворённость",
    "gap": "Gap",
    "fitbase_value": "Ценность Fitbase",
    "previous_solution_problems": "Проблемы прошлого решения",
    "drivers": "Драйверы",
    "barriers": "Барьеры",
    "lpr_quote": "Цитата ЛПР",
    "interview_label": "Интервью / ЛПР",
    "business_profile": "Профиль бизнеса",
    "lpr_profile": "Профиль ЛПР",
    "notes": "Заметки",
    "id": "ID",
}

# ===== XLSX =====

wb = Workbook()
ws = wb.active
ws.title = "Джобы (raw)"

# Стили
thin = Side(border_style="thin", color="CCCCCC")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_fill = PatternFill("solid", fgColor="2F4858")  # тёмно-серый
header_font = Font(bold=True, color="FFFFFF", size=11)
big_fill = PatternFill("solid", fgColor="D5F0D5")     # светло-зелёный
middle_fill = PatternFill("solid", fgColor="FFF4D5")  # светло-жёлтый
small_fill = PatternFill("solid", fgColor="F0F0F5")   # светло-серый
cell_align = Alignment(wrap_text=True, vertical="top")

# Заголовок
for c, key in enumerate(cols, 1):
    cell = ws.cell(row=1, column=c, value=HEADERS_RU.get(key, key))
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Данные
def clean_br(v):
    return (str(v or "")
            .replace("<br>", "\n").replace("<BR>", "\n")
            .replace("<br/>", "\n").replace("<br />", "\n"))

for r_idx, rec in enumerate(rows, 2):
    size = rec.get("job_size", "")
    row_fill = None
    if size == "Big":
        row_fill = big_fill
    elif size == "Middle":
        row_fill = middle_fill
    elif size == "Small":
        row_fill = small_fill
    for c, key in enumerate(cols, 1):
        cell = ws.cell(row=r_idx, column=c, value=clean_br(rec.get(key, "")))
        cell.border = border
        cell.alignment = cell_align
        if row_fill:
            cell.fill = row_fill

# Ширины столбцов
widths = {
    "job_size": 10, "business_segment": 14, "job_segment": 14, "want_short": 55,
    "context_when": 50, "want_result": 50, "so_that": 50,
    "importance": 14, "satisfaction": 16, "gap": 8, "fitbase_value": 40,
    "previous_solution_problems": 40, "drivers": 30, "barriers": 30, "lpr_quote": 50,
    "interview_label": 32, "business_profile": 65, "lpr_profile": 50, "notes": 30,
    "id": 8,
}
for c, key in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(c)].width = widths.get(key, 20)

# Freeze pane и автофильтр
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

# Вкладка README в начале
ws3 = wb.create_sheet("README", 0)
notes = [
    ["Таблица джобов Fitbase — единый свод из всех интервью"],
    [""],
    [f"Всего уникальных джобов: {len(rows)}"],
    ["Дубликаты (105 строк из 25 кластеров — итерации одних и тех же интервью в разных папках) удалены."],
    [""],
    ["Источники:"],
    ["1. 19 CSV из Interview_Analysis/segments/ (структурированные данные с привязкой к интервью)"],
    ["2. jobs_and_pains/jobs.csv — 29 агрегированных джобов с УТП (без привязки к конкретному интервью)"],
    ["3. analysis_catalina_jtbd.md — Каталина (эталонный AJTBD)"],
    ["4. Knowledge/Interviews/Padel/denis_padel_jtbd.md — Денис, падел"],
    ["5. Knowledge/Interviews/sales_calls_1c_switch/*.md — 5 интервью переходов с 1С"],
    ["6. Narrative MD (UrbanFit, Kochkin) — ссылки-плейсхолдеры (требуют ручной обработки)"],
    [""],
    ["Цветовая разметка:"],
    ["• Зелёный фон — Big Job"],
    ["• Жёлтый фон — Middle Job"],
    ["• Серый фон — Small Job"],
    [""],
    ["Сегменты:"],
    ["• S1=Студии | S2=Клубы одиночки | S3=Сетевые малые | S4=Сетевые крупные | S5=Без персонала | S6=Падел"],
    ["• J1=Владелец-оператор | J2=С админами | J3=Растущая студия | J4=Собственник сети | J5=Дети/семья | J6=Усталый ЛПР"],
]
for r_idx, row in enumerate(notes, 1):
    cell = ws3.cell(row=r_idx, column=1, value=row[0])
    if r_idx == 1:
        cell.font = Font(bold=True, size=14)
    elif row[0].endswith(":"):
        cell.font = Font(bold=True)
ws3.column_dimensions["A"].width = 100

wb.save(XLSX)
print(f"✅ XLSX: {XLSX}")

# ===== HTML =====

# Расшифровка сегментов — для подсказок и легенды
S_NAMES = {
    "S1": "Студии",
    "S2": "Клубы одиночки",
    "S3": "Сетевые малые (2–5 точек)",
    "S4": "Сетевые крупные (5+ точек)",
    "S5": "Без персонала",
    "S6": "Падел-корты",
}
J_NAMES = {
    "J1": "Владелец-оператор",
    "J2": "С администраторами",
    "J3": "Растущая студия",
    "J4": "Собственник сети",
    "J5": "Дети / семейные форматы",
    "J6": "Усталый ЛПР",
}

def label_segment(code, mapping):
    if not code or code == "?":
        return ""
    name = mapping.get(code, "")
    return f"{code} — {name}" if name else code


html_rows = []
for r in rows:
    size = r.get("job_size", "")
    cls = ""
    if size == "Big":
        cls = "row-big"
    elif size == "Middle":
        cls = "row-middle"
    elif size == "Small":
        cls = "row-small"
    # Заменяем краткие коды сегментов в самой ячейке на «S1 — Студии»
    r = dict(r)
    r["business_segment"] = label_segment(r.get("business_segment", ""), S_NAMES)
    r["job_segment"] = label_segment(r.get("job_segment", ""), J_NAMES)
    html_rows.append((cls, r))

def esc(s):
    if not s: return ""
    # Сначала нормализуем литералы <br> в реальные переносы строк
    s = str(s).replace("<br>", "\n").replace("<BR>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    s = (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
           .replace('"', "&quot;"))
    # И превращаем переносы в реальные <br> в HTML
    s = s.replace("\n", "<br>")
    return s

table_rows = []
for cls, r in html_rows:
    cells = "".join(f"<td>{esc(r.get(c,''))}</td>" for c in cols)
    table_rows.append(f'<tr class="{cls}">{cells}</tr>')

html = f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>Fitbase — таблица джобов</title>
<link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
<link rel="stylesheet" href="https://cdn.datatables.net/buttons/2.4.2/css/buttons.dataTables.min.css">
<style>
body {{ font-family: -apple-system, "SF Pro", "Helvetica Neue", Arial, sans-serif; margin: 20px; color: #1a1a1a; background: #f7f7f9; }}
h1 {{ font-size: 22px; margin-bottom: 4px; }}
.subtitle {{ color: #666; font-size: 13px; margin-bottom: 16px; }}
.legend {{ display: flex; gap: 14px; font-size: 12px; margin-bottom: 12px; flex-wrap: wrap; }}
.legend span {{ padding: 4px 10px; border-radius: 4px; }}
.l-big {{ background: #d5f0d5; }}
.l-middle {{ background: #fff4d5; }}
.l-small {{ background: #f0f0f5; }}
.l-dup {{ background: #f0d5d5; }}
table.dataTable {{ font-size: 12px; background: #fff; }}
table.dataTable thead th {{ background: #2f4858; color: #fff; padding: 10px 8px; }}
table.dataTable tbody td {{ padding: 8px 6px; vertical-align: top; max-width: 320px; }}
table.dataTable tbody td:nth-child({cols.index('business_profile')+1}),
table.dataTable thead th:nth-child({cols.index('business_profile')+1}) {{ min-width: 640px; max-width: 640px; }}
table.dataTable tbody td:nth-child({cols.index('lpr_profile')+1}),
table.dataTable thead th:nth-child({cols.index('lpr_profile')+1}) {{ min-width: 420px; max-width: 420px; }}
tr.row-big td {{ background: #e8f8e8; }}
tr.row-middle td {{ background: #fffae5; }}
tr.row-small td {{ background: #f5f5f9; }}
.dataTables_filter input {{ padding: 6px 10px; border: 1px solid #ccc; border-radius: 4px; }}
.summary {{ background: #fff; padding: 14px 18px; border-radius: 8px; margin: 14px 0; font-size: 13px; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }}
.summary b {{ color: #2f4858; }}
.warn {{ color: #c14040; }}
</style>
</head>
<body>

<h1>Fitbase — единая таблица джобов из всех интервью</h1>
<div class="subtitle">Дата сборки: автоматически. Источники: 19 CSV + 7 MD JTBD + 1 агрегированный CSV.</div>

<div class="summary">
<b>Всего уникальных джобов:</b> {len(rows)}.<br>
<b>Big:</b> {sum(1 for c,_ in html_rows if c == 'row-big')} • <b>Middle:</b> {sum(1 for c,_ in html_rows if c == 'row-middle')} • <b>Small:</b> {sum(1 for c,_ in html_rows if c == 'row-small')}<br>
Источники: 19 структурированных интервью + 7 MD JTBD-анализов + 29 агрегированных джобов с УТП.
</div>

<div class="legend">
<span class="l-big">Big Job</span>
<span class="l-middle">Middle Job</span>
<span class="l-small">Small Job</span>
</div>

<details style="background:#fff;padding:10px 14px;border-radius:8px;margin-bottom:14px;font-size:13px;box-shadow:0 1px 3px rgba(0,0,0,0.05);">
<summary style="cursor:pointer;font-weight:600;color:#2f4858;">Расшифровка сегментов (S и J)</summary>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:20px;margin-top:12px;">
<div>
<b style="color:#2f4858;">Бизнес-сегменты (S):</b><br>
<b>S1</b> — Студии (одна точка, моно-направление)<br>
<b>S2</b> — Клубы одиночки (один полноформатный клуб)<br>
<b>S3</b> — Сетевые малые (2–5 точек)<br>
<b>S4</b> — Сетевые крупные (5+ точек)<br>
<b>S5</b> — Без персонала (автономный зал)<br>
<b>S6</b> — Падел-корты
</div>
<div>
<b style="color:#2f4858;">Джоб-сегменты по Замесину (J):</b><br>
<b>J1</b> — Владелец-оператор (один за всё)<br>
<b>J2</b> — Руководитель с админами (делегирование)<br>
<b>J3</b> — Растущая студия (масштабирование)<br>
<b>J4</b> — Собственник сети (управление по показателям)<br>
<b>J5</b> — Детские / семейные форматы<br>
<b>J6</b> — Стабильный уставший ЛПР
</div>
</div>
</details>

<table id="jobs" class="display" style="width:100%">
<thead>
<tr>{"".join(f"<th>{HEADERS_RU.get(c,c)}</th>" for c in cols)}</tr>
</thead>
<tbody>
{chr(10).join(table_rows)}
</tbody>
</table>

<script src="https://code.jquery.com/jquery-3.7.0.min.js"></script>
<script src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
<script>
// Кастомная сортировка размеров: Big → Middle → Small → (пусто)
const SIZE_ORDER = {{ "Big": 1, "Middle": 2, "Small": 3, "": 9 }};
$.fn.dataTable.ext.type.order['job-size-pre'] = function(d) {{
  return SIZE_ORDER[d] !== undefined ? SIZE_ORDER[d] : 9;
}};

$(function() {{
  $('#jobs thead th').eq({cols.index('job_size')}).attr('data-type', 'job-size');
  $('#jobs').DataTable({{
    columnDefs: [{{ type: 'job-size', targets: {cols.index('job_size')} }}],
    pageLength: 50,
    lengthMenu: [25, 50, 100, 250, -1],
    order: [[ {cols.index('job_size')}, 'asc' ], [ {cols.index('business_segment')}, 'asc' ]],
    language: {{
      search: "Поиск:",
      lengthMenu: "Показать _MENU_ строк",
      info: "Показано _START_–_END_ из _TOTAL_",
      paginate: {{ previous: "←", next: "→" }},
      zeroRecords: "Ничего не найдено",
    }},
    fixedHeader: true,
    scrollX: true,
    deferRender: true,
  }});
}});
</script>

</body></html>
"""

HTML.write_text(html, encoding="utf-8")
print(f"✅ HTML: {HTML}")
