#!/usr/bin/env python3
"""
Экспорт jobs_clustered.csv в xlsx и интерактивный HTML с фокусом на:
- Размер → Блок → Канон. название → Частотность → Сегменты → so_that → ...
- Сортировка по частотности (приоритет для продакта)
- Фильтры по блоку и сегменту
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

OUT = Path("/Users/romansemenov/Desktop/Cloud Cod/Knowledge/jobs_master_table")
CSV_IN = OUT / "jobs_clustered.csv"
XLSX = OUT / "jobs_clustered.xlsx"
HTML = OUT / "clustered.html"

with open(CSV_IN, encoding="utf-8") as f:
    rows = list(csv.DictReader(f))

# Колонки в нужном порядке для отображения
DISPLAY_COLS = [
    "job_size",          # Размер
    "block",             # Блок (топ-уровень)
    "canonical_name",    # Каноническое название работы
    "frequency",         # Частотность
    "business_segments", # Сегменты (S1, S2...)
    "so_that",           # Чтобы
    "context_when",      # Когда
    "fitbase_value",     # Ценность Fitbase
    "previous_solution_problems",
    "drivers",
    "barriers",
    "lpr_quote",
    "sources_count",
    "sources",
    "all_variants",      # Все вариантные формулировки (для проверки)
]

HEADERS_RU = {
    "job_size": "Размер",
    "block": "Блок",
    "canonical_name": "Работа",
    "frequency": "Частотность",
    "business_segments": "Бизнес-сегменты",
    "job_segments": "Джоб-сегменты",
    "so_that": "Чтобы (цель/эффект)",
    "context_when": "Когда (контекст)",
    "fitbase_value": "Ценность Fitbase",
    "previous_solution_problems": "Проблемы прошлого решения",
    "drivers": "Драйверы",
    "barriers": "Барьеры",
    "lpr_quote": "Цитаты ЛПР (топ-3 длинных)",
    "sources_count": "Кол-во источников",
    "sources": "Список источников (интервью)",
    "all_variants": "Все исходные формулировки",
}

# Сортировка: размер Big→Middle→Small, потом по частотности desc
size_order = {"Big": 1, "Middle": 2, "Small": 3, "": 9}
rows.sort(key=lambda r: (size_order.get(r["job_size"], 9), -int(r["frequency"] or 0)))

# ====== XLSX ======
wb = Workbook()
ws = wb.active
ws.title = "Кластеры джобов"

thin = Side(border_style="thin", color="CCCCCC")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_fill = PatternFill("solid", fgColor="2F4858")
header_font = Font(bold=True, color="FFFFFF", size=11)
big_fill = PatternFill("solid", fgColor="D5F0D5")
middle_fill = PatternFill("solid", fgColor="FFF4D5")
small_fill = PatternFill("solid", fgColor="F0F0F5")
freq_bold = Font(bold=True, color="2F4858", size=11)
cell_align = Alignment(wrap_text=True, vertical="top")

# Заголовок
for c, key in enumerate(DISPLAY_COLS, 1):
    cell = ws.cell(row=1, column=c, value=HEADERS_RU.get(key, key))
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def clean_br(v):
    return (str(v or "")
            .replace("<br>", "\n").replace("<BR>", "\n")
            .replace("<br/>", "\n").replace("<br />", "\n"))

for r_idx, rec in enumerate(rows, 2):
    size = rec.get("job_size", "")
    row_fill = None
    if size == "Big": row_fill = big_fill
    elif size == "Middle": row_fill = middle_fill
    elif size == "Small": row_fill = small_fill
    for c, key in enumerate(DISPLAY_COLS, 1):
        val = rec.get(key, "")
        if key == "frequency":
            val = int(val) if val else 0
        cell = ws.cell(row=r_idx, column=c, value=clean_br(val) if key != "frequency" else val)
        cell.border = border
        cell.alignment = cell_align
        if row_fill:
            cell.fill = row_fill
        if key == "frequency":
            cell.font = freq_bold
            cell.alignment = Alignment(horizontal="center", vertical="top")

widths = {
    "job_size": 10, "block": 26, "canonical_name": 55,
    "frequency": 12, "business_segments": 18, "so_that": 45, "context_when": 45,
    "fitbase_value": 35, "previous_solution_problems": 35, "drivers": 25, "barriers": 25,
    "lpr_quote": 50, "sources_count": 8, "sources": 50, "all_variants": 60,
}
for c, key in enumerate(DISPLAY_COLS, 1):
    ws.column_dimensions[get_column_letter(c)].width = widths.get(key, 20)
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(DISPLAY_COLS))}{len(rows)+1}"

# README вкладка
ws_readme = wb.create_sheet("README", 0)
notes = [
    ["Кластеризованная таблица джобов Fitbase"],
    [""],
    [f"Из {len(rows) + 63} сырых джобов агрегировано в {len(rows)} уникальных работ"],
    ["(дубликаты внутри одного интервью + однотипные джобы между интервью объединены)"],
    [""],
    ["12 топ-блоков (таксономия из CLAUDE.md + новый блок «Личная эффективность владельца»):"],
    ["1. Финансы и учёт"],
    ["2. Персонал и зарплаты"],
    ["3. Клиенты и удержание"],
    ["4. Расписание и запись"],
    ["5. Коммуникации и маркетинг"],
    ["6. Аналитика и управление"],
    ["7. СКУД и доступ"],
    ["8. Операционка / не быть узким местом"],
    ["9. Запуск и переход"],
    ["10. Клиентский опыт и приложение"],
    ["11. Интеграции и техническое"],
    ["12. ✨ Личная эффективность владельца (новый, по запросу Стаса/StretchHouse)"],
    [""],
    ["Колонка «Частотность» — сколько разных интервью упомянули эту работу."],
    ["Колонка «Бизнес-сегменты» — в каких сегментах S1-S6 встречается."],
    ["Колонка «Все исходные формулировки» — конкретные вариации фраз из интервью (для проверки)."],
    [""],
    ["Сегменты:"],
    ["• S1=Студии | S2=Клубы одиночки | S3=Сетевые малые | S4=Сетевые крупные | S5=Без персонала | S6=Падел"],
]
for i, row in enumerate(notes, 1):
    cell = ws_readme.cell(row=i, column=1, value=row[0])
    if i == 1:
        cell.font = Font(bold=True, size=14)
ws_readme.column_dimensions["A"].width = 100

wb.save(XLSX)
print(f"✅ XLSX: {XLSX}")

# ====== HTML ======
S_NAMES = {"S1": "Студии", "S2": "Клубы одиночки", "S3": "Сетевые малые", "S4": "Сетевые крупные", "S5": "Без персонала", "S6": "Падел"}

def esc(s):
    if not s: return ""
    s = str(s).replace("<br>", "\n").replace("<BR>", "\n").replace("<br/>", "\n")
    s = (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;"))
    return s.replace("\n", "<br>")

def label_segments(codes):
    if not codes: return ""
    out = []
    for c in codes.split(","):
        c = c.strip()
        if c in S_NAMES:
            out.append(f"<span class='seg'>{c} — {S_NAMES[c]}</span>")
        elif c:
            out.append(f"<span class='seg'>{c}</span>")
    return " ".join(out)

table_rows = []
for r in rows:
    cls = ""
    s = r.get("job_size", "")
    if s == "Big": cls = "row-big"
    elif s == "Middle": cls = "row-middle"
    elif s == "Small": cls = "row-small"
    freq = int(r.get("frequency", 0) or 0)
    freq_class = "freq-hi" if freq >= 4 else ("freq-mid" if freq >= 2 else "freq-low")
    cells = [
        f"<td>{esc(r.get('job_size',''))}</td>",
        f"<td class='block-cell'>{esc(r.get('block',''))}</td>",
        f"<td class='work'><strong>{esc(r.get('canonical_name',''))}</strong></td>",
        f"<td class='freq {freq_class}'>{freq}</td>",
        f"<td>{label_segments(r.get('business_segments',''))}</td>",
        f"<td>{esc(r.get('so_that',''))}</td>",
        f"<td>{esc(r.get('context_when',''))}</td>",
        f"<td>{esc(r.get('fitbase_value',''))}</td>",
        f"<td>{esc(r.get('previous_solution_problems',''))}</td>",
        f"<td>{esc(r.get('lpr_quote',''))}</td>",
        f"<td>{esc(r.get('sources',''))}</td>",
        f"<td class='variants'>{esc(r.get('all_variants',''))}</td>",
    ]
    table_rows.append(f"<tr class='{cls}'>{''.join(cells)}</tr>")

headers_display = ["Размер", "Блок", "Работа", "Кол-во<br>интервью", "Сегменты", "Чтобы", "Когда", "Ценность Fitbase",
                   "Проблемы прошлого", "Цитаты ЛПР", "Источники", "Все исходные формулировки"]

html = f"""<!doctype html>
<html lang="ru">
<head>
<meta charset="utf-8">
<title>Fitbase — кластеризованные джобы</title>
<link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
<style>
body {{ font-family: -apple-system, "SF Pro", "Helvetica Neue", Arial, sans-serif; margin: 20px; color: #1a1a1a; background: #f7f7f9; }}
h1 {{ font-size: 22px; margin-bottom: 4px; }}
.subtitle {{ color: #666; font-size: 13px; margin-bottom: 16px; }}
.summary {{ background: #fff; padding: 14px 18px; border-radius: 8px; margin: 14px 0; font-size: 13px; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }}
.summary b {{ color: #2f4858; }}
table.dataTable {{ font-size: 12.5px; background: #fff; }}
table.dataTable thead th {{ background: #2f4858; color: #fff; padding: 10px 8px; font-weight: 600; }}
table.dataTable tbody td {{ padding: 10px 8px; vertical-align: top; max-width: 340px; }}
tr.row-big td {{ background: #e8f8e8; }}
tr.row-middle td {{ background: #fffae5; }}
tr.row-small td {{ background: #f5f5f9; }}
td.work {{ font-size: 13.5px; min-width: 280px; max-width: 360px; }}
td.work strong {{ color: #2f4858; }}
td.freq {{ text-align: center; font-weight: 700; font-size: 16px; min-width: 60px; }}
td.freq.freq-hi {{ color: #c14040; }}
td.freq.freq-mid {{ color: #d97a06; }}
td.freq.freq-low {{ color: #888; }}
td.block-cell {{ font-size: 11px; color: #555; min-width: 130px; }}
td.variants {{ color: #888; font-size: 11px; max-width: 280px; }}
.seg {{ display: inline-block; background: #e8f0f5; padding: 2px 6px; border-radius: 4px; margin: 1px; font-size: 11px; }}
.dataTables_filter input {{ padding: 6px 10px; border: 1px solid #ccc; border-radius: 4px; }}
.legend {{ display: flex; gap: 12px; font-size: 12px; margin-bottom: 12px; flex-wrap: wrap; }}
.legend span {{ padding: 4px 10px; border-radius: 4px; }}
.l-big {{ background: #d5f0d5; }}
.l-middle {{ background: #fff4d5; }}
.l-small {{ background: #f0f0f5; }}
.l-hi {{ background: #c14040; color: #fff; }}
.l-mid {{ background: #d97a06; color: #fff; }}
</style>
</head>
<body>
<h1>Fitbase — джобы клиентов, сгруппированные по работам</h1>
<div class="subtitle">{len(rows)} уникальных работ из 225 исходных джобов в 12 топ-блоках. Сортировка по размеру (Big→Middle→Small) и частотности.</div>

<div class="summary">
<b>Как читать:</b> «Частотность» = в скольких разных интервью встретилась эта работа.
Большая частотность + много сегментов = универсальная боль рынка.
Цвет цифры частотности: <span class="l-hi">≥4</span> <span class="l-mid">2-3</span> и обычный для уникальных.
</div>

<div class="legend">
<span class="l-big">Big Job</span>
<span class="l-middle">Middle Job</span>
<span class="l-small">Small Job</span>
</div>

<details style="background:#fff;padding:10px 14px;border-radius:8px;margin-bottom:14px;font-size:13px;box-shadow:0 1px 3px rgba(0,0,0,0.05);">
<summary style="cursor:pointer;font-weight:600;color:#2f4858;">12 блоков</summary>
<ol style="margin-top:8px">
<li>Финансы и учёт</li>
<li>Персонал и зарплаты</li>
<li>Клиенты и удержание</li>
<li>Расписание и запись</li>
<li>Коммуникации и маркетинг</li>
<li>Аналитика и управление</li>
<li>СКУД и доступ</li>
<li>Операционка / не быть узким местом</li>
<li>Запуск и переход</li>
<li>Клиентский опыт и приложение</li>
<li>Интеграции и техническое</li>
<li>✨ Личная эффективность владельца (новый блок)</li>
</ol>
</details>

<details style="background:#fff;padding:10px 14px;border-radius:8px;margin-bottom:14px;font-size:13px;box-shadow:0 1px 3px rgba(0,0,0,0.05);">
<summary style="cursor:pointer;font-weight:600;color:#2f4858;">Сегменты</summary>
<div style="margin-top:8px">
<b>S1</b> — Студии | <b>S2</b> — Клубы одиночки | <b>S3</b> — Сетевые малые (2-5 точек) |
<b>S4</b> — Сетевые крупные (5+) | <b>S5</b> — Без персонала | <b>S6</b> — Падел-корты
</div>
</details>

<table id="jobs" class="display" style="width:100%">
<thead>
<tr>{"".join(f"<th>{h}</th>" for h in headers_display)}</tr>
</thead>
<tbody>
{chr(10).join(table_rows)}
</tbody>
</table>

<script src="https://code.jquery.com/jquery-3.7.0.min.js"></script>
<script src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
<script>
const SIZE_ORDER = {{ "Big": 1, "Middle": 2, "Small": 3, "": 9 }};
$.fn.dataTable.ext.type.order['job-size-pre'] = d => SIZE_ORDER[d] !== undefined ? SIZE_ORDER[d] : 9;

$(function() {{
  $('#jobs').DataTable({{
    columnDefs: [
      {{ type: 'job-size', targets: 0 }},
      {{ type: 'num', targets: 3 }},
    ],
    pageLength: 50,
    lengthMenu: [25, 50, 100, 200, -1],
    order: [[ 0, 'asc' ], [ 3, 'desc' ]],
    language: {{
      search: "Поиск:",
      lengthMenu: "Показать _MENU_ строк",
      info: "Показано _START_–_END_ из _TOTAL_",
      paginate: {{ previous: "←", next: "→" }},
      zeroRecords: "Ничего не найдено",
    }},
    scrollX: true,
    deferRender: true,
  }});
}});
</script>
</body></html>
"""

HTML.write_text(html, encoding="utf-8")
print(f"✅ HTML: {HTML}")
